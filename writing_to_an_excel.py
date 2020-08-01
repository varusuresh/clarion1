import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

l1 = {'valuation_percentage': '100', 'cptyFixedThresholdType': 'fixed', 'Counter Party Minimum Transfer Amount Currency': 'NA', 'cptytriggertype': 'Fixed Values', 'gracePeriod': '', 'Counterparty Fixed Rounding Amount': 'NA', 'enableInterestCalc': 'NA', 'fixedVal': 'TRUE', 'resolution_time': 'NA', 'compound_interest_calc': 'NA', 'Rounding -- recall': 'down', 'rehypothecation_flag': 'NA', 'frequency': 'DAILY', 'Notification Time': 'NA', 'Base Currency': 'NA', 'Principal Fixed Rounding Amount': 'NA', 'Rounding -- Delivery': 'up', 'Counter Party Minimum Transfer Amount Value': 'NA', 'holiday_centre': 'NA', 'cash_only': 'NA', 'negative_interest': 'NA', 'Rating Method': 'Issue', 'netting_Interest': 'Net', 'Principal Minimum Transfer Amount Currency': 'NA', 'calcRule': 'First of the Month', 'Principal Fixed Trigger Threshold Currency': 'NA', 'netting_MTM_And_IA': 'NET', 'principaltriggertype': 'Fixed Values', 'Principal Fixed Rounding Amount Currency': 'NA', 'Principal Fixed Trigger Threshold Value': '0', 'interest_settlement_date': 'Same Day', 'resolution_date': 'NA', 'valuation_agent': 'NA', 'Counterparty Fixed Threshold Currency': 'NA', 'principalFixedThresholdType': 'fixed', 'Counterparty Fixed Rounding Amount Currency': 'NA', 'Principal Minimum Transfer Amount Value': 'NA', 'settlementDateAbbriviated': 'NA', 'applicable_to_both': 'false', 'Counterparty Fixed Trigger Threshold Value': '0', 'assetName': 'NA', 'Principal Fixed Trigger Threshold Type': 'fixed'}

#Getting the length of the dictionary
print(len(l1))

#Creating 2 lists for storing key and values obtained
datapoint= []
data_values = []

for key in l1.keys():
    datapoint.append(key)
for value in l1.values():
    data_values.append(value)

print(datapoint)
print(data_values)

#Number of rows
for i in range(1, len(datapoint)+1):
    #Number of columns
    for j in range(1, 2):
        sheet.cell(row = i, column = j, value = datapoint[i-1])
        sheet.cell(row=i, column= j+1, value = data_values[i-1])

#Saving the excel to desktop
wb.save("//home/dell/Desktop/demo.xlsx")
